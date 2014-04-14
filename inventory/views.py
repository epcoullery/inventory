# -*- encoding: utf-8 -*-
from __future__ import unicode_literals
from datetime import date

from django.contrib import admin, messages
from django.db import connection
from django.db.models import F, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import DetailView, CreateView, ListView

from .forms import MovementForm
from .models import Material, Order, Quantity, Room, Storage, Movement, Person


def cond_redirect(request):
    if request.user.is_authenticated():
        return redirect('/')
    else:
        return admin.site.login(request)

def home(request):
    cur_orders = Order.objects.filter(receive_date__isnull=True)
    missing_mat = Material.objects.annotate(total_quant=Sum('quantity__quantity')
                                 ).filter(total_quant__lt=F('threshold')
                                 ).exclude(pk__in=cur_orders.values_list('pk', flat=True))

    cursor = connection.cursor()
    cursor.execute("SELECT Sum(price*quantity) FROM inventory_quantity")
    total = cursor.fetchone()[0]

    context = {
        'missing_mat': missing_mat,
        'cur_orders': cur_orders,
        'rooms': Room.objects.all().prefetch_related('storage_set'),
        'total': total,
        'mov_years': Movement.objects.dates('when', 'year'),
    }
    return admin.site.index(request, extra_context=context)


class StorageView(DetailView):
    model = Storage
    template_name = 'inventory/storage.html'
    render_format = 'html'

    def get_context_data(self, **kwargs):
        context = super(StorageView, self).get_context_data(**kwargs)
        context.update({
            'other_storages': self.object.room.storage_set.exclude(pk=self.object.pk),
            'quant_items': self.object.quantity_set.select_related('material').extra(
                select={'lower_name':'lower(inventory_material.description)'}).order_by('lower_name'),
        })
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.render_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.writer.excel import save_virtual_workbook
            wb = Workbook()
            ws = wb.get_active_sheet()
            ws.title = 'Inventaire'
            # Headers
            headers = ['Matériel (salle %s, armoire %s)' % (self.object.room.number, self.object.code), 'Quantité', 'Unité']
            for col_idx, header in enumerate(headers):
                ws.cell(row=0, column=col_idx).value = header
                ws.cell(row=0, column=col_idx).style.font.bold = True
            # Data
            for row_idx, tr in enumerate(context['quant_items'], start=1):
                ws.cell(row=row_idx, column=0).value = unicode(tr.material)
                ws.cell(row=row_idx, column=1).value = tr.quantity
                ws.cell(row=row_idx, column=2).value = tr.material.unit

            ws.cell(row=row_idx+2, column=0).value = "État au %s" % date.today()
            ws.cell(row=row_idx+2, column=0).style.font.italic = True
            ws.column_dimensions['A'].width = 60

            response = HttpResponse(save_virtual_workbook(wb), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=exportation_%s_%s.xlsx' % (
                self.object.code.replace(' ', '_'), date.strftime(date.today(), '%Y-%m-%d'))
            return response
        else:
            return super(StorageView, self).render_to_response(context, **response_kwargs)


class QuantityEditView(CreateView):
    model = Movement
    form_class = MovementForm

    def dispatch(self, request, *args, **kwargs):
        self.storage = get_object_or_404(Storage, pk=kwargs['pk'])
        return super(QuantityEditView, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        self.material = None
        if 'mat' in request.GET:
            self.material = get_object_or_404(Material, pk=request.GET.get('mat'))
        return super(QuantityEditView, self).get(request, *args, **kwargs)

    def get_initial(self):
        initial = super(QuantityEditView, self).get_initial()
        if self.request.method == 'GET':
            initial.update({
                'storage': self.storage,
                'op_plus': self.request.GET.get('op') == 'true',
                'typ': 'use',
            })
            if self.material:
                initial['material'] = self.material
            if initial['op_plus']:
                initial['typ'] = 'order'
            try:
                initial['author'] = Person.objects.get(user=self.request.user)
            except Person.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        context = super(QuantityEditView, self).get_context_data(**kwargs)
        context.update({
            'storage': self.storage,
        })
        return context

    def form_valid(self, form):
        self.object = form.save()
        if form.cleaned_data['op_plus']:
            messages.success(self.request, "Vous avez ajouté du matériel avec succès")
        else:
            messages.success(self.request, "Vous avez retiré du matériel avec succès")
        return HttpResponse(self.object.storage.get_absolute_url())


class MovementExport(ListView):
    model = Movement

    def get_queryset(self):
        return self.model.objects.filter(when__year=self.kwargs['year'])

    def render_to_response(self, context, **response_kwargs):
        from openpyxl import Workbook
        from openpyxl.writer.excel import save_virtual_workbook
        wb = Workbook()
        ws = wb.get_active_sheet()
        ws.title = 'Mouvements'
        # Headers
        headers = ['Qui', 'Quand', 'Quoi', 'Où', 'Combien', 'Commentaire']
        for col_idx, header in enumerate(headers):
            ws.cell(row=0, column=col_idx).value = header
            ws.cell(row=0, column=col_idx).style.font.bold = True
        # Data
        for row_idx, tr in enumerate(self.object_list, start=1):
            ws.cell(row=row_idx, column=0).value = unicode(tr.author)
            ws.cell(row=row_idx, column=1).value = tr.when
            ws.cell(row=row_idx, column=2).value = unicode(tr.material)
            ws.cell(row=row_idx, column=3).value = unicode(tr.storage)
            ws.cell(row=row_idx, column=4).value = tr.quantity
            ws.cell(row=row_idx, column=5).value = tr.comment

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=mouvements_%s.xlsx' % (
            self.kwargs['year'])
        return response
